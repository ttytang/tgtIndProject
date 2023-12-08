#!/usr/bin/env python3
# coding=utf-8
#带有用户配置文件的调用方式：配置信息依靠读取用户配置文件
#说明：本地调用方式 python calculation.py user_config.xlsx 或者 python calculation.py
#      网络调用模式 python3 calculation.py user_config.xlsx kaoshi.xlsx pingshi.xlsx target.xlsx
#      未来支持没有显式的user_config.xlsx
#不带用户配置文件的调用方式：配置信息依靠搜索（方法待实现，接口保持一致）

from openpyxl import load_workbook
from openpyxl import Workbook
import sys
import xlaxop, re

class AchivInd:

    def __init__(self):
        self.achiIndInfoDict = {}
        self.testScoreFilePos = 'B7'
        self.usualScoreFilePos = 'B9'
        self.testIdRngPos = 'B8'
        self.usualIdRngPos = 'B10'
        self.rmvIdPos = 'B11'
        self.targetStartPos = 'A13'
        self.factorStartPos = 'B12'
        self.targetCnt = 0#initial value
        self.factorCnt = 0#initial value
        self.attributeCnt = 3#reference to user_config.xlsx for 3 attributes: 依赖编号、依赖占比、坐标信息
        self.targetBenchMarkScoreDict = {}
        self.targetClassScoreDict = {}
        self.targetIndictorDict = {}

    def readConfigInfo(self, configFile):
        if len(sys.argv)>=2:
            cfgFile = sys.argv[1]
        else:
            cfgFile = configFile
        wb = load_workbook(filename = cfgFile)
        self.confInfo = wb.active
    
    def getTestScoreFile(self):
        argvLen = len(sys.argv)
        if argvLen < 4:#local call with standalone mode
            testScoreFile = self.confInfo[self.testScoreFilePos].value#reference to user_config.xlsx
        elif argvLen >= 4:#remote call with network mode
            testScoreFile = sys.argv[2]#reference to user_config.xlsx
        return testScoreFile
        
    def getUsualScoreFile(self):
        argvLen = len(sys.argv)
        if argvLen < 4:#local call with standalone mode
            usualScoreFile = self.confInfo[self.usualScoreFilePos].value#reference to user_config.xlsx
        elif argvLen >= 4:#remote call with network mode
            usualScoreFile = sys.argv[3]#reference to user_config.xlsx
        return usualScoreFile
            
    def getTestIdRng(self):
        rngInfo = self.confInfo[self.testIdRngPos]#reference to user_cofig.xlsx
        return rngInfo.value.split('/')
        
    def getUsualIdRng(self):
        rngInfo = self.confInfo[self.usualIdRngPos]#reference to user_cofig.xlsx
        return rngInfo.value.split('/')
        
    def getRmvId(self):
        rmvIdInfo = self.confInfo[self.rmvIdPos]#reference to user_config.xlsx
        return rmvIdInfo.value.split('/')
    
    def __build1stLayerDict(self):#{target1:{},target2:{},...}
        targetCnt = 0
        targetPos = self.targetStartPos
        inCell = self.confInfo[targetPos]
        while (inCell.value[:2] == '目标'):#reference to user_cofig.xlsx marked with red
            targetCnt += 1
            self.achiIndInfoDict['target'+str(targetCnt)] = {}
            targetPos = targetPos[0]+str(int(targetPos[1:])+1)
            inCell = self.confInfo[targetPos]
        self.targetCnt = targetCnt
            
    def __build2ndLayerDict(self):#{target1:{factor1:{source:'kaoshi.xlsx',weight:0.7},factor2:{source:'pingshi.xlsx',weight:0.7}}, target2:{...}, ...}
        factorCnt = 0
        factorPos = self.factorStartPos
        inCell = self.confInfo[factorPos]
        while (inCell.value != None):
            if '考试' in inCell.value:#reference to user_cofig.xlsx marked with red
                sourceFlag = 1 #self.getTestScoreFile()
            else:
                sourceFlag = 0 #self.getUsualScoreFile()
            factorCnt += 1
            for firstDictInd in range(self.targetCnt):
                weightVal = self.confInfo[factorPos[0]+str(int(factorPos[1:])+firstDictInd+1)].value
                if (weightVal!=None) and (weightVal!='0'):
                    self.achiIndInfoDict['target'+str(firstDictInd+1)]['factor'+str(factorCnt)] = {}
                    self.achiIndInfoDict['target'+str(firstDictInd+1)]['factor'+str(factorCnt)]['source'] = sourceFlag
                    self.achiIndInfoDict['target'+str(firstDictInd+1)]['factor'+str(factorCnt)]['weight'] = weightVal                    
            #factorPos = chr(ord(factorPos[0])+1)+factorPos[1:]#didn't consider characeter above Z, such as AA, AB, ...
            factorPos = xlaxop.colInc(factorPos, 1)
            inCell = self.confInfo[factorPos]
        self.factorCnt = factorCnt

    def __build3rdLayerDict(self):#{target1:{factor1:{source:'kaoshi.xlsx',weight:0.7},factor2:{source:'pingshi.xlsx',weight:0.7}}, target2:{...}, ...}
        #charStartInd = self.factorStartPos[0]#didn't consider characeter above Z, such as AA, AB, ...
        #numStartInd = self.targetStartPos[1:]
        m1=re.match(r'^([A-Z]+)(\d+)$', self.factorStartPos)
        m2=re.match(r'^([A-Z]+)(\d+)$', self.targetStartPos)
        charStartInd = m1.group(1)
        numStartInd = m2.group(2)
        numInd = int(numStartInd)+len(self.achiIndInfoDict)+1
        #sorted(self.achiIndInfoDict)#Perhaps needed if default is not sorted
        for firstLayerKey in self.achiIndInfoDict:
            #sorted(self.achiIndInfoDict[firstLayerKey])#Perhaps needed if default is not sorted
            for secondLayerKey in self.achiIndInfoDict[firstLayerKey]:
                charInd = charStartInd
                for k in range(self.attributeCnt):
                    numIndToStr = str(numInd)
                    dpdtAttributeVal = self.confInfo[charInd+numIndToStr].value
                    if k == 1:#to support such as 3,3,5 instead of 11 in attribute2 fields
                        dpdtAttributeVal = str(dpdtAttributeVal)
                        dpdtAttributeVal = dpdtAttributeVal.replace('，',',').split(',')
                        dpdtAttributeVal = sum([int(float(i)) for i in dpdtAttributeVal])
                    self.achiIndInfoDict[firstLayerKey][secondLayerKey]['attribute'+str(k+1)]=dpdtAttributeVal#k=0:dpdtInd(the index list dependent to), k=1:dpdtPct(the percentage(X100) dependent to), k=2:dpdtPos(#the position dependent to)
                    #charInd = chr(ord(charInd)+1)#didn't consider characeter above Z, such as AA, AB, ...
                    charInd = xlaxop.colAlphaInc(charInd, 1)
                numInd += 1
            numInd += 1
            
    def buildAchiIndInfoDict(self):
        self.__build1stLayerDict()
        self.__build2ndLayerDict()
        self.__build3rdLayerDict()

    def __targetBenchMarkScore(self):
        for target in self.achiIndInfoDict:
            self.targetBenchMarkScoreDict[target] = 0
            for factor in self.achiIndInfoDict[target]:
                self.targetBenchMarkScoreDict[target] += self.achiIndInfoDict[target][factor]['weight']*self.achiIndInfoDict[target][factor]['attribute2']
        
    def __classScoreCal(self, wsSrc, verRng, horInds, excludeIds):
        scoreSum = 0
        idCnt = 0
        curPos = verRng[0]
        #beyondPos = verRng[1][0]+str(int(verRng[1][1:])+1)#assuming only A-Z column supported and exclued AA, AB, ....
        beyondPos = xlaxop.rowInc(verRng[1], 1)
        while curPos != beyondPos:
            if wsSrc[curPos].value not in excludeIds:
                idCnt += 1
                for j in horInds:
                    score = wsSrc[j].value
                    if score == None:#read out None if blank
                        score = 0
                    scoreSum += score
            #horInds = [ind[0]+str(int(ind[1:])+1) for ind in horInds]
            horInds = [xlaxop.rowInc(ind, 1) for ind in horInds]
            #curPos = curPos[0]+str(int(curPos[1:])+1)
            curPos = xlaxop.rowInc(curPos, 1)
        return scoreSum/idCnt#average of the score compared with benchmark score
    
    def __horIndsGen(self, horIndsStart, horIndsStr):
        horIndsPos = []
        horIndsStr = str(horIndsStr)
        horIndsStr = horIndsStr.replace('，',',').split(',')
        startPos = int(horIndsStr[0])
        for i in horIndsStr:
            #horIndsPos.append(chr(ord(horIndsStart[0])+int(i)-startPos)+horIndsStart[1:])#didn't support such as AA, AB, ....
            horIndsPos.append(xlaxop.colInc(horIndsStart,int(i)-startPos))
        return horIndsPos
    
    def __targetClassScore(self):
        wsTestscore = load_workbook(filename = self.getTestScoreFile()).active
        wsUsualscore = load_workbook(filename = self.getUsualScoreFile()).active
        excludeIds = self.getRmvId()
        for target in self.achiIndInfoDict:
            self.targetClassScoreDict[target] = 0
            for factor in self.achiIndInfoDict[target]:
                if self.achiIndInfoDict[target][factor]['source']==1: #1: sourceFlag is TestScoreFile; 
                    wsSrc = wsTestscore
                    verRng = self.getTestIdRng()
                else:
                    wsSrc = wsUsualscore #0: sourceFlag is UsualScoreFile
                    verRng = self.getUsualIdRng()
                horIndsStart = self.achiIndInfoDict[target][factor]['attribute3']
                horIndsStr = self.achiIndInfoDict[target][factor]['attribute1']
                horInds = self.__horIndsGen(horIndsStart, horIndsStr)
                factorClassScore = self.__classScoreCal(wsSrc, verRng, horInds, excludeIds)
                if (self.achiIndInfoDict[target][factor]['source']!=1):#usual score need to be rectified based on the times of homework
                    factorClassScore = factorClassScore/len(horInds)/100*self.achiIndInfoDict[target][factor]['attribute2']
                self.targetClassScoreDict[target] += self.achiIndInfoDict[target][factor]['weight']*factorClassScore

    def calTargetIndictor(self):
        self.__targetBenchMarkScore()
        self.__targetClassScore()
        for target in self.achiIndInfoDict:
            self.targetIndictorDict[target]=self.targetClassScoreDict[target]/self.targetBenchMarkScoreDict[target]

    def targetFileGen(self, file):
        workbook = Workbook()
        ws = workbook.active
        ws.title = '达成度结果'
        for i in range(1,3):#fixed value 2: 1, 2
            for j in range(1,len(self.achiIndInfoDict)+1):
                if i == 1:
                    ws.cell(row=i, column=j).value='目标'+str(j)
                else:
                    ws.cell(row=i, column=j).value=self.targetIndictorDict['target'+str(j)]
        workbook.save(filename=file)
        

if __name__=='__main__':
    achivInd = AchivInd()
    achivInd.readConfigInfo('./user_config.xlsx')
    #print(achivInd.getTestScoreFile())
    #print(achivInd.getUsualScoreFile())
    #print(achivInd.getTestIdRng())
    #print(achivInd.getUsualIdRng())
    #print(achivInd.getRmvId())
    #achivInd.build1stLayerDict()
    #print(achivInd.achiIndInfoDict)
    #achivInd.build2ndLayerDict()
    #print(achivInd.achiIndInfoDict)
    #achivInd.build3rdLayerDict()
    #print(achivInd.achiIndInfoDict)
    achivInd.buildAchiIndInfoDict()
    #print(achivInd.achiIndInfoDict)
    #achivInd.targetBenchMarkScore()
    #print(achivInd.targetBenchMarkScoreDict)
    #achivInd.targetClassScore()
    #print(achivInd.targetClassScoreDict)
    achivInd.calTargetIndictor()
    print(achivInd.targetIndictorDict)
    if len(sys.argv)==5:
        achivInd.targetFileGen(sys.argv[4])
    
    
    
        